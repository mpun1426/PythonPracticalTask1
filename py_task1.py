import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'https://www.jara.co.jp/member_list/'
html = requests.get(url)
soup = BeautifulSoup(html.content, 'html.parser')
prefectures = []
data = []
num = 1

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "スクレイピング課題1"

sheet["A1"].value = "都道府県"
sheet["B1"].value = "企業名"
sheet["C1"].value = "所在地"
sheet["D1"].value = "電話番号"
sheet.column_dimensions['A'].width = 10
sheet.column_dimensions['B'].width = 55
sheet.column_dimensions['C'].width = 50
sheet.column_dimensions['D'].width = 15

for div in soup.select("div.article_innner"):
    prefecture = div.find("h3").text
    sheet.cell(row = num + 1, column = 1, value = prefecture)
    prefectures.append(prefecture)

    for table in div.select(".listMap"):
        data_by_prefecture = table.find_all("tr")

        for tr in data_by_prefecture:
            company_data = tr.find_all("td")
            sheet.cell(row = num + 1, column = 2, value = company_data[0].text)
            sheet.cell(row = num + 1, column = 3, value = company_data[1].text)
            sheet.cell(row = num + 1, column = 4, value = company_data[2].text)
            num += 1
            data.append(company_data)

        num += 1

wb.save("py_task1.xlsx")
wb.close()
